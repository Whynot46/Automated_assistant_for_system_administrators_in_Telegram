import telebot
from telebot import types
from telebot.types import ReplyKeyboardRemove
from openpyxl import load_workbook
from datetime import datetime
import os
from markups import *
from IDs import *
from sys import platform


#очистка терминала
try:
    if platform == "linux" or platform == "linux2":
        os.system('clear')
    elif platform == "win32":
        os.system('cls')
except: pass


#Подключение к таблицам
try:
    wb_orders = load_workbook('./db/orders.xlsx')
    orders_sheet = wb_orders['orders']
    wb_users = load_workbook('./db/users.xlsx')
    users_sheet = wb_users['users']
    wb_reviews = load_workbook('./db/reviews.xlsx')
    reviews_sheet = wb_reviews['reviews']
    print('Loading tables: ok')
except:
    print('Loading tables: ERROR')


#Запуск бота @P@Nachalovo_support_bot
try:
    global bot
    bot = telebot.TeleBot(API_KEY)
    bot_name = '@Nachalovo_support_bot'
    print(f'Launching the bot {bot_name} : ok')
except:
    print(f'Launching the bot {bot_name} : ERROR')


#Стартовое сообщение
@bot.message_handler(commands=['start'])
def start_message(message):
    global users_current_line
    bot.send_message(message.chat.id,
                     'Вас приветсвует бот техподдержки Приволжского района👾\n\n'
                     'Данный бот предназначен для удобной и быстрой подачи заявок на обслуживание техники вашей организации🏢\n'
                     'Вы можете отправить заявку на ремонт, техническое обслуживание или консультацию по любому вопросу, связанному с работой вашей техники💻\n',
                     reply_markup=main_markup)
    user_registration(message)


#Регистрация пользователя
def user_registration(message):
    global users_current_line, users_sheet

    users_current_line = int(users_sheet[f'H1'].value)+2

    for line in range(1, users_current_line):
        if str(message.chat.id)==str(users_sheet[f'A{line}'].value):
            users_current_line = line
        else: 
            users_sheet['H1'] = int(users_sheet[f'H1'].value)+1

    bot.send_message(message.chat.id, '❗ Регистрация пользователя ❗')
    user_id = message.chat.id
    users_sheet[f'A{users_current_line}'] = user_id
    wb_users.save('./db/users.xlsx')
    input_organization(message)


def input_organization(message):
    
    bot.send_message(message.chat.id, '🏢 Укажите вашу организацию', reply_markup = organization_markup)
    bot.register_next_step_handler(message, check_organization)


def check_organization(message):

    if ( (message.text[0]=='/') or (message.text==None) ):
        bot.send_message(message.chat.id, "Недопустимое значение")
        input_organization(message)

    else: input_username(message)


def input_username(message):

    organization = message.text
    users_sheet[f'B{users_current_line}'] = organization
    wb_users.save('./db/users.xlsx')

    bot.send_message(message.chat.id, "🪪 Укажите ваше ФИО", reply_markup = ReplyKeyboardRemove())
    bot.register_next_step_handler(message, check_username)


def check_username(message):

    if ( (message.text[0]=='/') or (message.text==None) ):
        bot.send_message(message.chat.id, "Недопустимое значение")
        input_username(message)

    else: input_post(message)


def input_post(message):

    username = message.text
    users_sheet[f'C{users_current_line}'] = username
    wb_users.save('./db/users.xlsx')

    bot.send_message(message.chat.id, "📈 Укажите вашу должность")
    bot.register_next_step_handler(message, check_post)


def check_post(message):

    if ( (message.text[0]=='/') or (message.text==None) ):
        bot.send_message(message.chat.id, "Недопустимое значение")
        input_post(message)

    else: input_phone_number(message)


def input_phone_number(message):

    post = message.text
    users_sheet[f'D{users_current_line}'] = post
    wb_users.save('./db/users.xlsx')

    bot.send_message(message.chat.id, "📞 Укажите ваш контактный телефон")
    bot.register_next_step_handler(message, check_phone_number)


def check_phone_number(message):

    if ( (message.text[0]=='/') or (message.text==None) ):
        bot.send_message(message.chat.id, "Недопустимое значение")
        input_phone_number(message)

    else:
        phone_number = message.text
        users_sheet[f'E{users_current_line}'] = phone_number
        wb_users.save('./db/users.xlsx')

        bot.send_message(message.chat.id, '✅Успешно✅\n'
            '🗂 Зарегистрирован новый пользователь 🗂\n'
            f"🏢 Организация:\n{users_sheet[f'B{users_current_line}'].value}\n"
            f"🪪 ФИО: {users_sheet[f'C{users_current_line}'].value}\n"
            f"📈 Должность: {users_sheet[f'D{users_current_line}'].value}\n"
            f"📞Контактный телефон: {users_sheet[f'E{users_current_line}'].value}", reply_markup = main_markup)
        
        bot.send_message(main_chanel_id,
            '🗂 Зарегистрирован новый пользователь 🗂\n'
            f"🏢 Организация:\n{users_sheet[f'B{users_current_line}'].value}\n"
            f"🪪 ФИО: {users_sheet[f'C{users_current_line}'].value}\n"
            f"📈 Должность: {users_sheet[f'D{users_current_line}'].value}\n"
            f"📞Контактный телефон: {users_sheet[f'E{users_current_line}'].value}")



#Команда /help
@bot.message_handler(commands=['help'])
def commands(message):
    bot.send_message(message.chat.id, 'HELP🥸⁉️', reply_markup=main_markup)


#Формирование новой заявки
@bot.message_handler(commands=['Новая_заявка🗣'])
def new_request(message):

    users_current_line = 0

    max_line = int(reviews_sheet[f'I1'].value)+2

    for line in range (1, max_line):
        if str(users_sheet[f'A{line}'].value)==str(message.chat.id):
            users_current_line = line
            break

    bot.send_message(message.chat.id, "🗒 Заполните заявку по нижеуказанной форме")

    orders_current_line = int(orders_sheet['O1'].value)+2
    orders_sheet[f'A{orders_current_line}'] = int(orders_sheet['O1'].value)+1
    orders_sheet[f'K{orders_current_line}'] = users_sheet[f'A{users_current_line}'].value
    orders_sheet[f'B{orders_current_line}'] = users_sheet[f'B{users_current_line}'].value
    orders_sheet[f'C{orders_current_line}'] = users_sheet[f'C{users_current_line}'].value
    orders_sheet[f'D{orders_current_line}'] = users_sheet[f'D{users_current_line}'].value
    orders_sheet['O1'] = orders_sheet[f'A{orders_current_line}'].value
    wb_orders.save('./db/orders.xlsx')

    bot.send_message(message.chat.id, "#⃣ Укажите номер кабинета", reply_markup = ReplyKeyboardRemove())
    bot.register_next_step_handler(message, input_user_cabinet)
    

def input_user_cabinet(message):
    cabinet = message.text
    orders_sheet[f'E{orders_current_line}'] = cabinet
    wb_orders.save('./db/orders.xlsx')

    bot.send_message(message.chat.id, "📝 Подробно опишите вашу проблему")
    bot.register_next_step_handler(message, input_user_problem)

def input_user_problem(message):

    user_problem = message.text
    orders_sheet[f'F{orders_current_line}'] = user_problem
    wb_orders.save('./db/orders.xlsx')

    file_markup = types.ReplyKeyboardMarkup(resize_keyboard = True)
    btn_no_file = types.KeyboardButton(text = '❌ Не прикреплять ❌')
    file_markup.add(btn_no_file)
    bot.send_message(message.chat.id, "🌄 Прикрепите фото или видео", reply_markup = file_markup)
    bot.register_next_step_handler(message, input_user_file)

def input_user_file(message):

    orders_sheet[f'L{orders_current_line}'] = 'Доставлено'
    orders_sheet[f'H{orders_current_line}'] = f'{datetime.now().strftime("%d-%m-%Y %H:%M")}'
    orders_sheet[f'G{orders_current_line}'] = 'Отсутствуют'
    wb_orders.save('./db/orders.xlsx')

    request_message = bot.send_message(main_chanel_id, f"🏢{orders_sheet[f'B{orders_current_line}'].value}\n"
        f"🪪{orders_sheet[f'C{orders_current_line}'].value}\n"
        f"📈{orders_sheet[f'D{orders_current_line}'].value}\n"
        f"#⃣Кабинет: {orders_sheet[f'E{orders_current_line}'].value}\n"
        f"📝Проблема:\n{orders_sheet[f'F{orders_current_line}'].value}\n"
        f"🏷️ID: {orders_sheet[f'A{orders_current_line}'].value}\n"
        f"📨Дата создания: {orders_sheet[f'H{orders_current_line}'].value}",
        reply_markup = task_markup)

    request_hash = hash(request_message.text)
    orders_sheet[f'M{orders_current_line}'] = str(request_hash)
    wb_orders.save('./db/orders.xlsx')

    if message.text != '❌ Не прикреплять ❌':
        orders_sheet[f'G{orders_current_line}'] = 'Прикреплены'
        wb_orders.save('./db/orders.xlsx')
        bot.forward_message(main_chanel_id, message.chat.id, message.id)

    bot.send_message(message.chat.id, "✅ Заявка успешно доставлена ✅", reply_markup = main_markup)
    print(f"New request received    | Request ID: {orders_sheet[f'A{orders_current_line}'].value} |")


#Просмотр заявок пользователя
@bot.message_handler(commands=['Мои_заявки🔍'])
def show_user_requests(message):

    max_line = int(orders_sheet[f'O1'].value)+2
    k=0
    for line in range (2, max_line):
        if str(orders_sheet[f'K{line}'].value)==str(users_sheet[f'A{users_current_line}'].value):
            k+=1
            bot.send_message(message.chat.id,
                f"#⃣Кабинет: {orders_sheet[f'E{line}'].value}\n"
                f"📝Проблема:\n{orders_sheet[f'F{line}'].value}\n"
                f"🏷️ID: {orders_sheet[f'A{line}'].value}\n"
                f"🆘Статус: {orders_sheet[f'L{line}'].value}\n"
                f"📨Дата создания: {orders_sheet[f'H{line}'].value}\n"
                f"👀Дата просмотра: {orders_sheet[f'I{line}'].value}\n"
                f"🤝Дата выполнения: {orders_sheet[f'J{line}'].value}",
                reply_markup = main_markup)

    if k==0: bot.send_message(message.chat.id, '⚠ Вы ещё не оставляли заявок ⚠', reply_markup = main_markup)


#Контакты админа
@bot.message_handler(commands=['Контакты📜'])
def show_contacts(message):
    bot.send_message(message.chat.id, '👨‍💻 Скляренко Екатерина Владимировна\n📱WhatsApp: +79171775068\n', reply_markup = main_markup)


#Отзывы
@bot.message_handler(commands=['Оставить_отзыв📫'])
def make_new_review(message):

    max_line = int(reviews_sheet[f'I1'].value)+2

    for line in range (1, max_line):
        if str(users_sheet[f'A{line}'].value)==str(message.chat.id):
            users_current_line = line
            break

    bot.send_message(message.chat.id, "Напишите всё, что думаете о данном сервисе🙃\nВаш отзыв будет учитываться при дальнейших обновлениях сервиса👾", reply_markup = ReplyKeyboardRemove())
    bot.register_next_step_handler(message, save_user_review)


def save_user_review(message):
    reviews_current_line = int(reviews_sheet['I1'].value)+2
    text_review = message.text
    reviews_sheet[f'A{reviews_current_line}'] = str(reviews_current_line-1)
    reviews_sheet[f'B{reviews_current_line}'] = users_sheet[f'B{users_current_line}'].value
    reviews_sheet[f'C{reviews_current_line}'] = users_sheet[f'C{users_current_line}'].value
    reviews_sheet[f'D{reviews_current_line}'] = users_sheet[f'D{users_current_line}'].value
    reviews_sheet[f'E{reviews_current_line}'] = text_review
    reviews_sheet['I1'] = reviews_sheet[f'A{reviews_current_line}'].value
    reviews_sheet[f'F{reviews_current_line}'] = f'{datetime.now().strftime("%d-%m-%Y %H:%M")}'
    wb_reviews.save('./db/reviews.xlsx')
    bot.send_message(main_chanel_id, '👀Получен новый отзыв👀\n'
        f"🏢{reviews_sheet[f'B{reviews_current_line}'].value}\n"
        f"🪪{reviews_sheet[f'C{reviews_current_line}'].value}\n"
        f"📈{reviews_sheet[f'D{reviews_current_line}'].value}\n"
        f"Отзыв:\n{reviews_sheet[f'E{reviews_current_line}'].value}\n"
        f"{reviews_sheet[f'F{reviews_current_line}'].value}")
    bot.send_message(message.chat.id, '💌 Ваш отзыв успешно отправлен!\nОгромное спасибо за обратную связь🥰' ,reply_markup = main_markup)
    print(f"New review received | from_user: {reviews_sheet[f'C{reviews_current_line}'].value} |")


#callback-и inline кнопок Принятия/Отказа
@bot.callback_query_handler(func = lambda call : True)
def answer(call):
    users_current_line = int(orders_sheet[f'O1'].value)+2
    if call.data == 'done':
        for line in range (2, users_current_line):
            if str(orders_sheet[f'M{line}'].value)== str(hash(call.message.text)):
                if str(orders_sheet[f'L{line}'].value)=='Доставлено':
                    bot.edit_message_text(f'{call.message.text}\n🧐Принято: {datetime.now().strftime("%d-%m-%Y %H:%M")}', reply_markup = lets_done_markup, chat_id=call.message.chat.id, message_id=call.message.message_id)
                    request_id = str(orders_sheet[f'A{line}'].value)
                    print(f'Request received   Requset ID: {request_id}')
                    orders_sheet[f'I{line}'] = str(datetime.now().strftime("%d-%m-%Y %H:%M"))
                    orders_sheet[f'L{line}'] = 'Принято'
                    wb_orders.save('./db/orders.xlsx')
                    bot.send_message(int(orders_sheet[f'K{line}'].value), f'☑️Ваша заявка принята в рассмотрение\n🏷️ID заявки: {request_id}')
                    break

    elif call.data == 'refusal':
        for line in range (2, users_current_line):
            if str(orders_sheet[f'M{line}'].value)== str(hash(call.message.text)):
                if str(orders_sheet[f'L{line}'].value)=='Доставлено':
                    bot.edit_message_text(f'{call.message.text}\nОтказано: {datetime.now().strftime("%d-%m-%Y %H:%M")}', reply_markup = refusal_markup, chat_id=call.message.chat.id, message_id=call.message.message_id)
                    request_id = str(orders_sheet[f'A{line}'].value)
                    print(f'Refusal of the request  | Requset ID: {request_id} |')
                    orders_sheet[f'J{line}'] = str(datetime.now().strftime("%d-%m-%Y %H:%M"))
                    orders_sheet[f'L{line}'] = 'Отказано'
                    wb_orders.save('./db/orders.xlsx')
                    bot.send_message(int(orders_sheet[f'K{line}'].value), f'❌Вашу заявку отклонили\n🏷️ID заявки: {request_id}')
                    break

    elif call.data == 'done_request':
        for line in range (2, users_current_line):
            if str(orders_sheet[f'M{line}'].value)== str(hash(call.message.text.partition('\n🧐Принято: ')[0])):
                if str(orders_sheet[f'L{line}'].value)=='Принято':
                    bot.edit_message_text(f'{call.message.text}\n✅Выполнено: {datetime.now().strftime("%d-%m-%Y %H:%M")}', reply_markup=done_markup, chat_id=call.message.chat.id, message_id=call.message.message_id)
                    request_id = str(orders_sheet[f'A{line}'].value)
                    print(f'Request completed   | Requset ID: {request_id} |')
                    orders_sheet[f'J{line}'] = str(datetime.now().strftime("%d-%m-%Y %H:%M"))
                    orders_sheet[f'L{line}'] = 'Выполнено'
                    wb_orders.save('./db/orders.xlsx')
                    bot.send_message(int(orders_sheet[f'K{line}'].value), f'✅Ваша заявка была выполнена\n🏷️ID заявки: {request_id}')
                    break

    elif call.data == 'pass':
        pass


#Панель администратора
@bot.message_handler(commands=['admin_panel'])
def open_admin_panel(message):
        if str(message.chat.id) in admin_id:
            bot.send_message(message.chat.id, 'Панель администратора', reply_markup = admin_markup)


#Выгрузка таблицы с пользователями
@bot.message_handler(commands=['get_users'])
def get_users_sheet(message):
    if str(message.chat.id) in admin_id:
        bot.send_document(message.chat.id, open(r'./db/users.xlsx', 'rb'))


#Выгрузка таблицы заявок
@bot.message_handler(commands=['get_orders'])
def get_orders_sheet(message):
    if str(message.chat.id) in admin_id:
        bot.send_document(message.chat.id, open(r'./db/orders.xlsx', 'rb'))


#Выгрузка таблицы отзывов
@bot.message_handler(commands=['get_reviews'])
def get_reviews_sheet(message):
    if str(message.chat.id) in admin_id:
        bot.send_document(message.chat.id, open(r'./db/reviews.xlsx', 'rb'))


#Общая рассылка
@bot.message_handler(commands=['tell_everyone'])
def send_message_for_everyone(message):
    if str(message.chat.id) in admin_id:
        bot.send_message(message.chat.id, "Укажите текст рассылки", reply_markup=cancel_markup)
        bot.register_next_step_handler(message, send_admin_message_for_everyone)


def send_admin_message_for_everyone(message):
    if str(message.text)!='Отмена':
        max_line = int(users_sheet[f'H1'].value)+2
        for line in range (2, max_line):
            try: 
                bot.send_message(int(users_sheet[f'A{line}'].value), message.text)
            except: 
                bot.send_message(message.chat.id, f"Ошибка при отправке сообщению пользователю с ID: {str(users_sheet[f'A{line}'].value)}")
                continue
        bot.send_message(message.chat.id, 'Сообщения успешно отправлены', reply_markup = main_markup)

    else: bot.send_message(message.chat.id, 'Отменено', reply_markup = main_markup)  


#Сообщение для конкретного пользователя
@bot.message_handler(commands=['tell_to_user'])
def send_message_for_user(message):
    if str(message.chat.id) in admin_id:
        bot.send_message(message.chat.id, "Укажите ID пользователя, которому вы хотите отправить сообщение", reply_markup=cancel_markup)
        bot.register_next_step_handler(message, check_recipient_id)


def check_recipient_id(message):
    global recipient_id
    recipient_id = int(message.text)
    print(recipient_id)
    if str(message.text)=='Отмена':
        bot.send_message(message.chat.id, 'Отменено', reply_markup = main_markup)  
    elif (str(message.text[0]) == '/') or (str(message.text[0]) == None):
        bot.send_message(message.chat.id, f'Недопустимое значение')
        send_message_for_user(message)
    else:
        bot.send_message(message.chat.id, "Укажите текст сообщения", reply_markup=cancel_markup)
        bot.register_next_step_handler(message, send_admin_message_for_recipient_user)


def send_admin_message_for_recipient_user(message):
    global recipient_id
    try: 
        bot.send_message(recipient_id, message.text)
        bot.send_message(message.chat.id, 'Сообщения успешно отправлены', reply_markup = main_markup)
    except: 
        bot.send_message(message.chat.id, f"Ошибка при отправке сообщению пользователю с ID: {str(recipient_id)}")



if __name__ == "__main__":
    #while True:
        #try:
    bot.polling(none_stop=True, interval=0, timeout=20)
        #except: pass
